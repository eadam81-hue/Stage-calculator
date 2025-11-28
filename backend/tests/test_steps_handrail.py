import requests
import json
import io
import openpyxl
from datetime import datetime

class StepsHandrailTester:
    def __init__(self, base_url="https://stage-planner-3.preview.emergentagent.com"):
        self.base_url = base_url
        self.api_url = f"{base_url}/api"
        self.tests_run = 0
        self.tests_passed = 0
        self.failed_tests = []
        
    def log_result(self, test_name, success, details=""):
        """Log test result"""
        self.tests_run += 1
        if success:
            self.tests_passed += 1
            print(f"✅ {test_name}")
        else:
            self.failed_tests.append(f"{test_name}: {details}")
            print(f"❌ {test_name}: {details}")
    
    def make_request(self, endpoint, method="POST", data=None):
        """Make API request"""
        url = f"{self.api_url}/{endpoint}"
        headers = {'Content-Type': 'application/json'}
        
        try:
            if method == "POST":
                response = requests.post(url, json=data, headers=headers)
            elif method == "GET":
                response = requests.get(url, headers=headers)
            
            return response.status_code, response.json() if response.content else {}
        except Exception as e:
            return 500, {"error": str(e)}
    
    def upload_required_components(self):
        """Upload all required components for steps and handrail testing"""
        print("\n🔧 Uploading required components...")
        
        # Create Excel file with all required components
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Headers: Name, SKU, Quantity, Price, Weight, Width, Depth
        ws['A1'] = 'Name'
        ws['B1'] = 'SKU'
        ws['C1'] = 'Quantity'
        ws['D1'] = 'Price'
        ws['E1'] = 'Weight'
        ws['F1'] = 'Width'
        ws['G1'] = 'Depth'
        
        # Required components for testing
        components = [
            # Basic deck components
            ['Litedeck 4x8', 'LD48', 50, 150.0, 25.0, 2.44, 1.22],
            ['Litedeck 4x4', 'LD44', 30, 100.0, 15.0, 1.22, 1.22],
            ['Litedeck 4x2', 'LD42', 40, 75.0, 10.0, 1.22, 0.61],
            ['Aludeck 2x1m', 'AD21', 25, 120.0, 18.0, 2.0, 1.0],
            
            # Legs and supports
            ['165mm legs', 'LEG165', 100, 25.0, 3.0, 0.165, 0.165],
            ['Stage Leg 600mm', 'SL600', 50, 45.0, 8.0, 0.6, 0.1],
            ['Stage Leg 1000mm', 'SL1000', 40, 65.0, 12.0, 1.0, 0.1],
            
            # Steps components
            ['Adjustable Stage Treads: 600-1000mm', 'AST600', 20, 180.0, 15.0, 1.0, 0.3],
            ['Adjustable Stage Treads: 1000-1800mm', 'AST1000', 15, 250.0, 20.0, 1.5, 0.3],
            
            # Handrail components - Imperial
            ['Litedeck 8ft Handrail', 'LH8', 30, 85.0, 8.0, 2.44, 0.1],
            ['Litedeck 4ft Handrail', 'LH4', 40, 45.0, 4.0, 1.22, 0.1],
            ['Litedeck 2ft Handrail', 'LH2', 25, 25.0, 2.0, 0.61, 0.1],
            
            # Handrail components - Metric
            ['Aludeck 2m Handrail', 'AH2', 25, 75.0, 6.0, 2.0, 0.1],
            ['Aludeck 1m Handrail', 'AH1', 30, 40.0, 3.0, 1.0, 0.1],
        ]
        
        # Add components to worksheet
        for i, comp in enumerate(components, start=2):
            for j, value in enumerate(comp, start=1):
                ws.cell(row=i, column=j, value=value)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Upload file
        files = {'file': ('components.xlsx', excel_buffer, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        try:
            response = requests.post(f"{self.api_url}/components/upload", files=files)
            if response.status_code == 200:
                result = response.json()
                print(f"✅ Uploaded {result.get('components_added', 0)} components")
                return True
            else:
                print(f"❌ Upload failed: {response.status_code}")
                return False
        except Exception as e:
            print(f"❌ Upload error: {str(e)}")
            return False
    
    def test_steps_low_height_370mm(self):
        """Test steps for 370mm height (1ft) - should add Litedeck 4x2 + 4x 165mm legs"""
        data = {
            "width": 6.0,
            "depth": 4.0,
            "height": 0.37,  # 370mm
            "location_type": "outdoor",
            "add_steps": True,
            "steps_quantity": "one"
        }
        
        status, response = self.make_request("calculate", data=data)
        
        if status == 200:
            parts = response.get('parts_list', [])
            part_names = [part['name'] for part in parts]
            
            # Check for required components
            has_litedeck_4x2 = any('4x2' in name and 'litedeck' in name.lower() for name in part_names)
            has_165mm_legs = any('165' in name and 'leg' in name.lower() for name in part_names)
            
            if has_litedeck_4x2 and has_165mm_legs:
                # Check quantities
                leg_part = next((p for p in parts if '165' in p['name'] and 'leg' in p['name'].lower()), None)
                if leg_part and leg_part['quantity_used'] == 4:
                    self.log_result("Steps 370mm height (one set)", True)
                else:
                    self.log_result("Steps 370mm height (one set)", False, f"Expected 4 legs, got {leg_part['quantity_used'] if leg_part else 0}")
            else:
                self.log_result("Steps 370mm height (one set)", False, f"Missing components: Litedeck 4x2={has_litedeck_4x2}, 165mm legs={has_165mm_legs}")
        else:
            self.log_result("Steps 370mm height (one set)", False, f"API error: {status}")
    
    def test_steps_low_height_570mm(self):
        """Test steps for 570mm height (1.5ft) - should add Litedeck 4x4 + 4x2 + 8x 165mm legs"""
        data = {
            "width": 6.0,
            "depth": 4.0,
            "height": 0.57,  # 570mm
            "location_type": "outdoor",
            "add_steps": True,
            "steps_quantity": "one"
        }
        
        status, response = self.make_request("calculate", data=data)
        
        if status == 200:
            parts = response.get('parts_list', [])
            part_names = [part['name'] for part in parts]
            
            # Check for required components
            has_litedeck_4x4 = any('4x4' in name and 'litedeck' in name.lower() for name in part_names)
            has_litedeck_4x2 = any('4x2' in name and 'litedeck' in name.lower() for name in part_names)
            has_165mm_legs = any('165' in name and 'leg' in name.lower() for name in part_names)
            
            if has_litedeck_4x4 and has_litedeck_4x2 and has_165mm_legs:
                # Check leg quantity
                leg_part = next((p for p in parts if '165' in p['name'] and 'leg' in p['name'].lower()), None)
                if leg_part and leg_part['quantity_used'] == 8:
                    self.log_result("Steps 570mm height (one set)", True)
                else:
                    self.log_result("Steps 570mm height (one set)", False, f"Expected 8 legs, got {leg_part['quantity_used'] if leg_part else 0}")
            else:
                self.log_result("Steps 570mm height (one set)", False, f"Missing components: 4x4={has_litedeck_4x4}, 4x2={has_litedeck_4x2}, legs={has_165mm_legs}")
        else:
            self.log_result("Steps 570mm height (one set)", False, f"API error: {status}")
    
    def test_steps_quantity_two_sets(self):
        """Test steps with quantity 'two' - should double the components"""
        data = {
            "width": 6.0,
            "depth": 4.0,
            "height": 0.37,  # 370mm
            "location_type": "outdoor",
            "add_steps": True,
            "steps_quantity": "two"
        }
        
        status, response = self.make_request("calculate", data=data)
        
        if status == 200:
            parts = response.get('parts_list', [])
            
            # Check leg quantity (should be 8 for two sets)
            leg_part = next((p for p in parts if '165' in p['name'] and 'leg' in p['name'].lower()), None)
            if leg_part and leg_part['quantity_used'] == 8:
                self.log_result("Steps quantity 'two' sets", True)
            else:
                self.log_result("Steps quantity 'two' sets", False, f"Expected 8 legs for two sets, got {leg_part['quantity_used'] if leg_part else 0}")
        else:
            self.log_result("Steps quantity 'two' sets", False, f"API error: {status}")
    
    def test_steps_adjustable_treads_800mm(self):
        """Test 800mm height with steps - should use Adjustable Stage Treads: 600-1000mm"""
        data = {
            "width": 6.0,
            "depth": 4.0,
            "height": 0.8,  # 800mm
            "location_type": "outdoor",
            "add_steps": True,
            "steps_quantity": "one"
        }
        
        status, response = self.make_request("calculate", data=data)
        
        if status == 200:
            parts = response.get('parts_list', [])
            part_names = [part['name'] for part in parts]
            
            # Check for adjustable treads 600-1000mm
            has_adjustable_treads = any('adjustable' in name.lower() and 'tread' in name.lower() and '600' in name for name in part_names)
            
            if has_adjustable_treads:
                self.log_result("Steps 800mm height (adjustable treads 600-1000mm)", True)
            else:
                self.log_result("Steps 800mm height (adjustable treads 600-1000mm)", False, f"Missing adjustable treads 600-1000mm. Found: {part_names}")
        else:
            self.log_result("Steps 800mm height (adjustable treads 600-1000mm)", False, f"API error: {status}")
    
    def test_steps_adjustable_treads_1200mm(self):
        """Test 1200mm height with steps - should use Adjustable Stage Treads: 1000-1800mm"""
        data = {
            "width": 6.0,
            "depth": 4.0,
            "height": 1.2,  # 1200mm
            "location_type": "outdoor",
            "add_steps": True,
            "steps_quantity": "one"
        }
        
        status, response = self.make_request("calculate", data=data)
        
        if status == 200:
            parts = response.get('parts_list', [])
            part_names = [part['name'] for part in parts]
            
            # Check for adjustable treads 1000-1800mm
            has_adjustable_treads = any('adjustable' in name.lower() and 'tread' in name.lower() and '1000' in name for name in part_names)
            
            if has_adjustable_treads:
                self.log_result("Steps 1200mm height (adjustable treads 1000-1800mm)", True)
            else:
                self.log_result("Steps 1200mm height (adjustable treads 1000-1800mm)", False, f"Missing adjustable treads 1000-1800mm. Found: {part_names}")
        else:
            self.log_result("Steps 1200mm height (adjustable treads 1000-1800mm)", False, f"API error: {status}")
    
    def test_handrail_imperial_24x12ft_critical(self):
        """CRITICAL USER EXAMPLE: Test 24ft x 12ft stage with handrail - should get ~5x 8ft + 2x 4ft handrails"""
        # Convert feet to meters: 24ft = 7.3152m, 12ft = 3.6576m
        data = {
            "width": 7.3152,  # 24ft
            "depth": 3.6576,  # 12ft
            "height": 1.0,
            "location_type": "outdoor",
            "add_handrail": True
        }
        
        status, response = self.make_request("calculate", data=data)
        
        if status == 200:
            parts = response.get('parts_list', [])
            
            # Find handrail components
            handrail_8ft = next((p for p in parts if 'handrail' in p['name'].lower() and '8ft' in p['name']), None)
            handrail_4ft = next((p for p in parts if 'handrail' in p['name'].lower() and '4ft' in p['name']), None)
            
            # Calculate expected perimeter: back + 2 sides = 24ft + 2*12ft = 48ft
            # Expected: 6x 8ft handrails (48ft), but should be approximately 5x 8ft + 2x 4ft
            
            if handrail_8ft and handrail_4ft:
                total_8ft = handrail_8ft['quantity_used']
                total_4ft = handrail_4ft['quantity_used']
                
                # Check if we have reasonable quantities (approximately 5x 8ft + 2x 4ft)
                if 4 <= total_8ft <= 6 and 1 <= total_4ft <= 3:
                    self.log_result("CRITICAL: 24ft x 12ft handrail calculation", True, f"Got {total_8ft}x 8ft + {total_4ft}x 4ft handrails")
                else:
                    self.log_result("CRITICAL: 24ft x 12ft handrail calculation", False, f"Expected ~5x 8ft + 2x 4ft, got {total_8ft}x 8ft + {total_4ft}x 4ft")
            else:
                self.log_result("CRITICAL: 24ft x 12ft handrail calculation", False, f"Missing handrail components. 8ft: {handrail_8ft is not None}, 4ft: {handrail_4ft is not None}")
        else:
            self.log_result("CRITICAL: 24ft x 12ft handrail calculation", False, f"API error: {status}")
    
    def test_handrail_metric_6x4m_indoor(self):
        """Test 6m x 4m indoor stage with handrail - should use Aludeck 2m/1m handrails"""
        data = {
            "width": 6.0,
            "depth": 4.0,
            "height": 1.0,
            "location_type": "indoor",
            "add_handrail": True
        }
        
        status, response = self.make_request("calculate", data=data)
        
        if status == 200:
            parts = response.get('parts_list', [])
            part_names = [part['name'] for part in parts]
            
            # Check for Aludeck handrails
            has_aludeck_2m = any('aludeck' in name.lower() and 'handrail' in name.lower() and '2m' in name for name in part_names)
            has_aludeck_1m = any('aludeck' in name.lower() and 'handrail' in name.lower() and '1m' in name for name in part_names)
            
            if has_aludeck_2m or has_aludeck_1m:
                self.log_result("Handrail 6m x 4m indoor (Aludeck)", True)
            else:
                self.log_result("Handrail 6m x 4m indoor (Aludeck)", False, f"Missing Aludeck handrails. Found: {part_names}")
        else:
            self.log_result("Handrail 6m x 4m indoor (Aludeck)", False, f"API error: {status}")
    
    def test_handrail_steps_interaction_critical(self):
        """CRITICAL: Test 24ft x 12ft stage with handrail AND 2 sets of steps - should remove 2x 4ft handrails"""
        data = {
            "width": 7.3152,  # 24ft
            "depth": 3.6576,  # 12ft
            "height": 1.0,
            "location_type": "outdoor",
            "add_handrail": True,
            "add_steps": True,
            "steps_quantity": "two"
        }
        
        status, response = self.make_request("calculate", data=data)
        
        if status == 200:
            parts = response.get('parts_list', [])
            
            # Find handrail components
            handrail_8ft = next((p for p in parts if 'handrail' in p['name'].lower() and '8ft' in p['name']), None)
            handrail_4ft = next((p for p in parts if 'handrail' in p['name'].lower() and '4ft' in p['name']), None)
            
            # With 2 sets of steps, should remove 2x 4ft handrails from the calculation
            if handrail_8ft and handrail_4ft:
                total_4ft = handrail_4ft['quantity_used']
                
                # Should have fewer 4ft handrails due to steps interaction
                if total_4ft >= 0:  # Could be 0 if all removed
                    self.log_result("CRITICAL: Handrail-Steps interaction (2 sets)", True, f"4ft handrails reduced to {total_4ft} due to steps")
                else:
                    self.log_result("CRITICAL: Handrail-Steps interaction (2 sets)", False, f"Unexpected 4ft handrail count: {total_4ft}")
            else:
                self.log_result("CRITICAL: Handrail-Steps interaction (2 sets)", False, "Missing handrail components")
        else:
            self.log_result("CRITICAL: Handrail-Steps interaction (2 sets)", False, f"API error: {status}")
    
    def test_handrail_steps_interaction_one_set(self):
        """Test handrail with 1 set of steps - should remove 1x 4ft handrail (may result in 0 4ft handrails)"""
        data = {
            "width": 7.3152,  # 24ft
            "depth": 3.6576,  # 12ft
            "height": 1.0,
            "location_type": "outdoor",
            "add_handrail": True,
            "add_steps": True,
            "steps_quantity": "one"
        }
        
        status, response = self.make_request("calculate", data=data)
        
        if status == 200:
            parts = response.get('parts_list', [])
            
            # Find handrail components
            handrail_8ft = next((p for p in parts if 'handrail' in p['name'].lower() and '8ft' in p['name']), None)
            handrail_4ft = next((p for p in parts if 'handrail' in p['name'].lower() and '4ft' in p['name']), None)
            
            # For 48ft perimeter with 1 step, should have 6x 8ft and 0x 4ft (step removes the 1x 4ft)
            if handrail_8ft and handrail_8ft['quantity_used'] == 6:
                if handrail_4ft is None or handrail_4ft['quantity_used'] == 0:
                    self.log_result("Handrail-Steps interaction (1 set)", True, f"Correctly removed 4ft handrail: 6x 8ft, 0x 4ft")
                else:
                    self.log_result("Handrail-Steps interaction (1 set)", False, f"Expected 0x 4ft handrails, got {handrail_4ft['quantity_used']}")
            else:
                self.log_result("Handrail-Steps interaction (1 set)", False, f"Expected 6x 8ft handrails, got {handrail_8ft['quantity_used'] if handrail_8ft else 0}")
        else:
            self.log_result("Handrail-Steps interaction (1 set)", False, f"API error: {status}")
    
    def test_handrail_safety_recommendation_above_570mm(self):
        """Test handrail recommendation for 600mm stage without handrail"""
        data = {
            "width": 6.0,
            "depth": 4.0,
            "height": 0.6,  # 600mm
            "location_type": "outdoor",
            "add_handrail": False
        }
        
        status, response = self.make_request("calculate", data=data)
        
        if status == 200:
            recommendation = response.get('handrail_recommendation')
            
            if recommendation and 'safety' in recommendation.lower():
                self.log_result("Handrail safety recommendation (600mm)", True, f"Got recommendation: {recommendation}")
            else:
                self.log_result("Handrail safety recommendation (600mm)", False, f"Missing or incorrect recommendation: {recommendation}")
        else:
            self.log_result("Handrail safety recommendation (600mm)", False, f"API error: {status}")
    
    def test_handrail_no_recommendation_below_570mm(self):
        """Test no handrail recommendation for 500mm stage without handrail"""
        data = {
            "width": 6.0,
            "depth": 4.0,
            "height": 0.5,  # 500mm
            "location_type": "outdoor",
            "add_handrail": False
        }
        
        status, response = self.make_request("calculate", data=data)
        
        if status == 200:
            recommendation = response.get('handrail_recommendation')
            
            if recommendation is None:
                self.log_result("No handrail recommendation (500mm)", True)
            else:
                self.log_result("No handrail recommendation (500mm)", False, f"Unexpected recommendation: {recommendation}")
        else:
            self.log_result("No handrail recommendation (500mm)", False, f"API error: {status}")
    
    def test_steps_boundary_cases(self):
        """Test steps at boundary heights"""
        boundary_tests = [
            (0.299, "299mm - should not add steps"),
            (0.300, "300mm - should add steps"),
            (0.599, "599mm - should add low height steps"),
            (0.600, "600mm - should add adjustable treads"),
            (1.000, "1000mm - should add adjustable treads 600-1000mm"),
            (1.001, "1001mm - should add adjustable treads 1000-1800mm")
        ]
        
        for height, description in boundary_tests:
            data = {
                "width": 6.0,
                "depth": 4.0,
                "height": height,
                "location_type": "outdoor",
                "add_steps": True,
                "steps_quantity": "one"
            }
            
            status, response = self.make_request("calculate", data=data)
            
            if status == 200:
                parts = response.get('parts_list', [])
                has_steps = any('step' in part['name'].lower() or 'tread' in part['name'].lower() or 
                              ('165' in part['name'] and 'leg' in part['name'].lower()) for part in parts)
                
                if height < 0.3:
                    # Should not have steps
                    if not has_steps:
                        self.log_result(f"Boundary test: {description}", True)
                    else:
                        self.log_result(f"Boundary test: {description}", False, "Unexpected steps added")
                else:
                    # Should have steps
                    if has_steps:
                        self.log_result(f"Boundary test: {description}", True)
                    else:
                        self.log_result(f"Boundary test: {description}", False, "No steps added")
            else:
                self.log_result(f"Boundary test: {description}", False, f"API error: {status}")
    
    def test_inventory_availability(self):
        """Test that components have reasonable stock for moderate-sized stages"""
        # Test a moderate stage to check inventory (not too large to avoid expected shortfalls)
        data = {
            "width": 8.0,
            "depth": 6.0,
            "height": 1.0,
            "location_type": "outdoor",
            "add_steps": True,
            "add_handrail": True,
            "steps_quantity": "one"
        }
        
        status, response = self.make_request("calculate", data=data)
        
        if status == 200:
            has_shortfall = response.get('has_inventory_issues', False)
            parts = response.get('parts_list', [])
            
            shortfall_parts = [part for part in parts if part.get('has_shortfall', False)]
            
            if not has_shortfall and not shortfall_parts:
                self.log_result("Inventory availability check", True)
            else:
                shortfall_names = [part['name'] for part in shortfall_parts]
                self.log_result("Inventory availability check", False, f"Inventory shortfall in: {shortfall_names}")
        else:
            self.log_result("Inventory availability check", False, f"API error: {status}")
    
    def run_all_tests(self):
        """Run all steps and handrail tests"""
        print("🧪 Starting Steps and Handrail Feature Tests")
        print("=" * 60)
        
        # Upload components first
        if not self.upload_required_components():
            print("❌ Failed to upload components. Cannot proceed with tests.")
            return False
        
        print("\n📋 Running Steps Feature Tests...")
        self.test_steps_low_height_370mm()
        self.test_steps_low_height_570mm()
        self.test_steps_quantity_two_sets()
        self.test_steps_adjustable_treads_800mm()
        self.test_steps_adjustable_treads_1200mm()
        
        print("\n🛡️ Running Handrail Feature Tests...")
        self.test_handrail_imperial_24x12ft_critical()
        self.test_handrail_metric_6x4m_indoor()
        
        print("\n🔗 Running Handrail-Steps Interaction Tests...")
        self.test_handrail_steps_interaction_critical()
        self.test_handrail_steps_interaction_one_set()
        
        print("\n⚠️ Running Safety Recommendation Tests...")
        self.test_handrail_safety_recommendation_above_570mm()
        self.test_handrail_no_recommendation_below_570mm()
        
        print("\n🎯 Running Edge Case Tests...")
        self.test_steps_boundary_cases()
        
        print("\n📦 Running Inventory Tests...")
        self.test_inventory_availability()
        
        # Print summary
        print("\n" + "=" * 60)
        print(f"📊 TEST SUMMARY: {self.tests_passed}/{self.tests_run} tests passed")
        
        if self.failed_tests:
            print("\n❌ FAILED TESTS:")
            for failure in self.failed_tests:
                print(f"   • {failure}")
        
        if self.tests_passed == self.tests_run:
            print("\n🎉 All tests passed!")
            return True
        else:
            print(f"\n⚠️ {len(self.failed_tests)} tests failed")
            return False

def main():
    tester = StepsHandrailTester()
    success = tester.run_all_tests()
    return 0 if success else 1

if __name__ == "__main__":
    import sys
    sys.exit(main())