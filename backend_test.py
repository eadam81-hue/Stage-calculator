import requests
import sys
import json
import io
from datetime import datetime
import openpyxl

class StageCalculatorAPITester:
    def __init__(self, base_url="https://stage-planner-2.preview.emergentagent.com"):
        self.base_url = base_url
        self.api_url = f"{base_url}/api"
        self.tests_run = 0
        self.tests_passed = 0
        self.component_ids = []

    def run_test(self, name, method, endpoint, expected_status, data=None, files=None):
        """Run a single API test"""
        url = f"{self.api_url}/{endpoint}"
        headers = {}
        
        if files is None:
            headers['Content-Type'] = 'application/json'

        self.tests_run += 1
        print(f"\n🔍 Testing {name}...")
        print(f"   URL: {url}")
        
        try:
            if method == 'GET':
                response = requests.get(url, headers=headers)
            elif method == 'POST':
                if files:
                    response = requests.post(url, files=files)
                else:
                    response = requests.post(url, json=data, headers=headers)
            elif method == 'DELETE':
                response = requests.delete(url, headers=headers)

            success = response.status_code == expected_status
            if success:
                self.tests_passed += 1
                print(f"✅ Passed - Status: {response.status_code}")
                try:
                    response_data = response.json()
                    print(f"   Response: {json.dumps(response_data, indent=2)[:200]}...")
                    return True, response_data
                except:
                    return True, {}
            else:
                print(f"❌ Failed - Expected {expected_status}, got {response.status_code}")
                try:
                    error_data = response.json()
                    print(f"   Error: {error_data}")
                except:
                    print(f"   Error: {response.text}")
                return False, {}

        except Exception as e:
            print(f"❌ Failed - Error: {str(e)}")
            return False, {}

    def create_test_excel_file(self):
        """Create a test Excel file with sample components"""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Headers
        ws['A1'] = 'Name'
        ws['B1'] = 'Quantity'
        ws['C1'] = 'Price'
        ws['D1'] = 'Weight'
        
        # Sample data
        test_data = [
            ['Stage Deck Panel', 20, 125.50, 45.2],
            ['Support Beam', 15, 89.99, 32.1],
            ['Frame Connector', 50, 12.75, 2.5],
            ['Outdoor Support Leg', 8, 156.00, 78.3],
            ['Platform Base', 10, 200.25, 95.7]
        ]
        
        for i, row_data in enumerate(test_data, start=2):
            for j, value in enumerate(row_data, start=1):
                ws.cell(row=i, column=j, value=value)
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer

    def test_root_endpoint(self):
        """Test the root API endpoint"""
        return self.run_test("Root Endpoint", "GET", "", 200)

    def test_upload_components(self):
        """Test uploading components via Excel file"""
        excel_file = self.create_test_excel_file()
        files = {'file': ('test_components.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        
        success, response = self.run_test(
            "Upload Components Excel",
            "POST",
            "components/upload",
            200,
            files=files
        )
        
        if success and response.get('components_added', 0) > 0:
            print(f"   Successfully added {response['components_added']} components")
            return True
        return False

    def test_get_components(self):
        """Test getting all components"""
        success, response = self.run_test(
            "Get All Components",
            "GET",
            "components",
            200
        )
        
        if success and isinstance(response, list):
            print(f"   Found {len(response)} components")
            # Store component IDs for later deletion tests
            self.component_ids = [comp['id'] for comp in response if 'id' in comp]
            return True
        return False

    def test_calculate_stage_indoor(self):
        """Test stage calculation for indoor installation"""
        calculation_data = {
            "width": 10.0,
            "depth": 8.0,
            "height": 1.5,
            "location_type": "indoor"
        }
        
        success, response = self.run_test(
            "Calculate Indoor Stage",
            "POST",
            "calculate",
            200,
            data=calculation_data
        )
        
        if success:
            required_fields = ['width', 'depth', 'height', 'location_type', 'parts_list', 'total_price', 'total_weight']
            missing_fields = [field for field in required_fields if field not in response]
            if missing_fields:
                print(f"   Warning: Missing fields in response: {missing_fields}")
                return False
            
            print(f"   Total Price: ${response.get('total_price', 0)}")
            print(f"   Total Weight: {response.get('total_weight', 0)} kg")
            print(f"   Parts Count: {len(response.get('parts_list', []))}")
            return True
        return False

    def test_calculate_stage_outdoor(self):
        """Test stage calculation for outdoor installation"""
        calculation_data = {
            "width": 12.0,
            "depth": 10.0,
            "height": 2.0,
            "location_type": "outdoor"
        }
        
        success, response = self.run_test(
            "Calculate Outdoor Stage",
            "POST",
            "calculate",
            200,
            data=calculation_data
        )
        
        if success:
            print(f"   Total Price: ${response.get('total_price', 0)}")
            print(f"   Total Weight: {response.get('total_weight', 0)} kg")
            print(f"   Parts Count: {len(response.get('parts_list', []))}")
            return True
        return False

    def test_get_calculations(self):
        """Test getting calculation history"""
        success, response = self.run_test(
            "Get Calculation History",
            "GET",
            "calculations",
            200
        )
        
        if success and isinstance(response, list):
            print(f"   Found {len(response)} calculations in history")
            return True
        return False

    def test_delete_component(self):
        """Test deleting a single component"""
        if not self.component_ids:
            print("   Skipping - No components available to delete")
            return True
            
        component_id = self.component_ids[0]
        success, response = self.run_test(
            "Delete Single Component",
            "DELETE",
            f"components/{component_id}",
            200
        )
        
        if success:
            self.component_ids.remove(component_id)
            return True
        return False

    def test_delete_all_components(self):
        """Test deleting all components"""
        success, response = self.run_test(
            "Delete All Components",
            "DELETE",
            "components",
            200
        )
        
        if success:
            print(f"   Deleted {response.get('deleted_count', 0)} components")
            self.component_ids.clear()
            return True
        return False

    def test_calculate_without_components(self):
        """Test calculation when no components are available"""
        calculation_data = {
            "width": 5.0,
            "depth": 4.0,
            "height": 1.0,
            "location_type": "indoor"
        }
        
        success, response = self.run_test(
            "Calculate Without Components",
            "POST",
            "calculate",
            400,  # Should return 400 when no components available
            data=calculation_data
        )
        
        return success

def main():
    print("🚀 Starting Stage Calculator API Tests")
    print("=" * 50)
    
    tester = StageCalculatorAPITester()
    
    # Test sequence
    test_results = []
    
    # Basic connectivity
    test_results.append(tester.test_root_endpoint())
    
    # Component management
    test_results.append(tester.test_upload_components())
    test_results.append(tester.test_get_components())
    
    # Stage calculations
    test_results.append(tester.test_calculate_stage_indoor())
    test_results.append(tester.test_calculate_stage_outdoor())
    
    # History
    test_results.append(tester.test_get_calculations())
    
    # Component deletion
    test_results.append(tester.test_delete_component())
    test_results.append(tester.test_delete_all_components())
    
    # Edge case - no components
    test_results.append(tester.test_calculate_without_components())
    
    # Print final results
    print("\n" + "=" * 50)
    print(f"📊 Test Results: {tester.tests_passed}/{tester.tests_run} passed")
    
    if tester.tests_passed == tester.tests_run:
        print("🎉 All tests passed!")
        return 0
    else:
        print("❌ Some tests failed")
        return 1

if __name__ == "__main__":
    sys.exit(main())